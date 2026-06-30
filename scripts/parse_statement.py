#!/usr/bin/env python3
import argparse
import csv
import json
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path


ALIASES = {
    "trade_date": ["成交日期", "交易日期", "发生日期", "委托日期", "日期", "business_date", "trade date", "date"],
    "security_code": ["证券代码", "股票代码", "代码", "合约代码", "证券编号", "stock_code", "symbol", "ticker"],
    "security_name": ["证券名称", "股票名称", "名称", "合约名称", "stock_name", "security_name"],
    "side": ["买卖方向", "交易方向", "操作", "买卖标志", "业务名称", "direction", "side", "action"],
    "price": ["成交价格", "成交均价", "价格", "成交价", "price", "fill_price"],
    "quantity": ["成交数量", "成交股数", "数量", "股数", "成交量", "quantity", "qty", "shares"],
    "trade_amount": ["成交金额", "成交额", "委托金额", "amount", "gross_amount", "turnover"],
    "commission": ["手续费", "佣金", "交易佣金", "commission", "fee"],
    "stamp_tax": ["印花税", "stamp_tax"],
    "transfer_fee": ["过户费", "transfer_fee"],
    "other_fee": ["其他费用", "规费", "经手费", "证管费", "结算费", "other_fee"],
    "cash_amount": ["发生金额", "总发生金额", "清算金额", "资金发生额", "资金变动", "net_amount", "cash_amount"],
    "cash_balance": ["资金余额", "可用余额", "余额", "资金结余", "cash_balance", "balance"],
}

OUTPUT_FIELDS = [
    "trade_date",
    "security_code",
    "security_name",
    "side",
    "price",
    "quantity",
    "trade_amount",
    "commission",
    "stamp_tax",
    "transfer_fee",
    "other_fee",
    "total_fee",
    "cash_amount",
    "cash_balance",
    "source_row",
]

REQUIRED = ["trade_date", "side", "price", "quantity"]


def norm(text):
    return re.sub(r"[\s_\-:：/()（）]+", "", str(text or "").strip().lower())


def read_csv(path):
    last_error = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            raw = Path(path).read_text(encoding=encoding)
            sample = raw[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
            except csv.Error:
                dialect = csv.excel
            rows = list(csv.DictReader(raw.splitlines(), dialect=dialect))
            return rows
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"无法读取 CSV: {last_error}")


def read_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 XLSX 需要安装 openpyxl") from exc
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return []
    headers = [str(v).strip() if v is not None else "" for v in values[0]]
    rows = []
    for row in values[1:]:
        rows.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})
    return rows


def read_statement(path):
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in (".xlsx", ".xlsm"):
        return read_xlsx(path)
    raise ValueError("仅支持 .csv、.xlsx、.xlsm")


def build_mapping(headers, manual_mapping=None):
    manual_mapping = manual_mapping or {}
    normalized_headers = {norm(h): h for h in headers}
    mapping = {}
    suggestions = {}

    for canonical, source in manual_mapping.items():
        if source in headers:
            mapping[canonical] = source

    for canonical, aliases in ALIASES.items():
        if canonical in mapping:
            continue
        candidates = [canonical] + aliases
        for candidate in candidates:
            source = normalized_headers.get(norm(candidate))
            if source is not None:
                mapping[canonical] = source
                break
        if canonical not in mapping:
            fuzzy = []
            for header in headers:
                nh = norm(header)
                if any(norm(alias) in nh or nh in norm(alias) for alias in aliases if norm(alias)):
                    fuzzy.append(header)
            suggestions[canonical] = fuzzy[:5]

    return mapping, suggestions


def parse_number(value, default=0.0):
    if value is None or value == "":
        return default
    text = str(value).strip().replace(",", "").replace("￥", "").replace("¥", "")
    text = text.replace("%", "")
    if text in ("--", "-", "无", "nan", "None"):
        return default
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    try:
        number = float(text)
        return -number if negative else number
    except ValueError:
        return default


def parse_date(value):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)) and value > 20000:
        return (datetime(1899, 12, 30) + timedelta(days=int(value))).date().isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def normalize_side(value):
    text = str(value or "").strip().upper()
    if any(token in text for token in ("买入", "买", "BUY", "PURCHASE", "B")):
        return "BUY"
    if any(token in text for token in ("卖出", "卖", "SELL", "REDEMPTION", "S")):
        return "SELL"
    return text


def get(row, mapping, canonical, default=""):
    source = mapping.get(canonical)
    return row.get(source, default) if source else default


def clean_rows(rows, mapping):
    cleaned = []
    for idx, row in enumerate(rows, start=2):
        price = abs(parse_number(get(row, mapping, "price")))
        quantity = abs(parse_number(get(row, mapping, "quantity")))
        trade_amount_raw = parse_number(get(row, mapping, "trade_amount"), default=None)
        trade_amount = abs(trade_amount_raw) if trade_amount_raw is not None else price * quantity
        commission = abs(parse_number(get(row, mapping, "commission")))
        stamp_tax = abs(parse_number(get(row, mapping, "stamp_tax")))
        transfer_fee = abs(parse_number(get(row, mapping, "transfer_fee")))
        other_fee = abs(parse_number(get(row, mapping, "other_fee")))
        total_fee = commission + stamp_tax + transfer_fee + other_fee
        side = normalize_side(get(row, mapping, "side"))
        cash_amount = parse_number(get(row, mapping, "cash_amount"), default=None)
        if cash_amount is None:
            cash_amount = -(trade_amount + total_fee) if side == "BUY" else trade_amount - total_fee
        cleaned.append({
            "trade_date": parse_date(get(row, mapping, "trade_date")),
            "security_code": str(get(row, mapping, "security_code")).strip(),
            "security_name": str(get(row, mapping, "security_name")).strip(),
            "side": side,
            "price": round(price, 6),
            "quantity": round(quantity, 6),
            "trade_amount": round(trade_amount, 6),
            "commission": round(commission, 6),
            "stamp_tax": round(stamp_tax, 6),
            "transfer_fee": round(transfer_fee, 6),
            "other_fee": round(other_fee, 6),
            "total_fee": round(total_fee, 6),
            "cash_amount": round(cash_amount, 6),
            "cash_balance": round(parse_number(get(row, mapping, "cash_balance"), default=0.0), 6),
            "source_row": idx,
        })
    cleaned.sort(key=lambda r: (r["trade_date"], r["source_row"]))
    return cleaned


def write_csv(rows, output):
    with open(output, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def load_manual_mapping(path):
    if not path:
        return {}
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def main():
    parser = argparse.ArgumentParser(description="Parse a broker statement into cleaned_trades.csv")
    parser.add_argument("input_file")
    parser.add_argument("-o", "--output", default="cleaned_trades.csv")
    parser.add_argument("--field-map", help="JSON mapping from canonical fields to source columns")
    parser.add_argument("--suggestions-out", default="field_mapping_suggestions.json")
    args = parser.parse_args()

    rows = read_statement(args.input_file)
    if not rows:
        raise SystemExit("输入文件没有数据行")
    headers = list(rows[0].keys())
    mapping, suggestions = build_mapping(headers, load_manual_mapping(args.field_map))
    missing = [field for field in REQUIRED if field not in mapping]
    if "security_code" not in mapping and "security_name" not in mapping:
        missing.append("security_code or security_name")
    if missing:
        payload = {"mapped_fields": mapping, "unmapped_required": missing, "suggestions": suggestions, "source_columns": headers}
        Path(args.suggestions_out).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        raise SystemExit(2)

    cleaned = clean_rows(rows, mapping)
    write_csv(cleaned, args.output)
    payload = {"mapped_fields": mapping, "unmapped_optional": {k: v for k, v in suggestions.items() if k not in mapping}}
    Path(args.suggestions_out).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {args.output} ({len(cleaned)} rows)")
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
